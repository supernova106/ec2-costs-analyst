#!/usr/bin/python
import argparse
import os
import json
import re
import copy
import math
import time
import random
import datetime
import sys
import collections
import boto
import boto.vpc
import boto.ec2
import boto.ec2.autoscale
import boto.ec2.elb
import boto.ec2.networkinterface
import boto.sns
import requests
import subprocess
import requests
from pprint import pprint
from boto.ec2.autoscale import AutoScaleConnection, LaunchConfiguration, ScalingPolicy
from boto.ec2.regioninfo import RegionInfo
from prettytable import PrettyTable
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter

LINUX_ON_DEMAND_PRICE_URL ='http://a0.awsstatic.com/pricing/1/ec2/linux-od.min.js'
LINUX_ON_DEMAND_PREVIOUS_GEN_PRICE_URL ='http://a0.awsstatic.com/pricing/1/ec2/previous-generation/linux-od.min.js'

region = os.getenv('REGION')

# You can uncomment and set these, or set the env variables AWSAccessKeyId & AWSSecretKey
# AWS_ACCESS_KEY_ID="aaaaaaaaaaaaaaaaaaaa"
# AWS_SECRET_ACCESS_KEY="bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb"

try:
	AWS_ACCESS_KEY_ID
except NameError:
	try:
		AWS_ACCESS_KEY_ID=os.environ['AWS_ACCESS_KEY_ID']
		AWS_SECRET_ACCESS_KEY=os.environ['AWS_SECRET_ACCESS_KEY']
	except KeyError:
		print "Please set env variable"
		sys.exit(1)

def get_price_table(url):
	"""Get and return price table
	"""
	resp = requests.get(url)
	content = resp.content
	callback_prefix = 'callback('
	callback_suffix = ');'
	prefix_index = content.find(callback_prefix) + len(callback_prefix)
	suffix_index = content.rfind(callback_suffix)
	content = content[prefix_index:suffix_index]
	# do a little regular expression hack to quote key name to make the
	# content becomes JSON format
	content = re.sub(r'(\w+?):', r'"\1":', content)
	return json.loads(content)

def price_table_to_price_mapping(table):
	"""Convert price table to a dict mapping from region to instance type
	to instance info
	"""
	region_price_mapping = {}
	for region_table in table['config']['regions']:
		types = {}
		for type_category in region_table['instanceTypes']:
			for size in type_category['sizes']:
				types[size['size']] = size
		region_price_mapping[region_table['region']] = types
	return region_price_mapping

if __name__ == "__main__":
	parser = argparse.ArgumentParser()
	parser.add_argument("--region", help='Input the region: us-east-1,us-west-2', required=True)
	parser.add_argument("--format", help='Output format: [xlsx]. None to print out to screen', required=False)
	args = vars(parser.parse_args())
	region = args['region']
	format = args['format']

	#ec2 connection
	ec2_conn = boto.ec2.connect_to_region(region)
	reservations = ec2_conn.get_all_instances()

	od_price_table = get_price_table(LINUX_ON_DEMAND_PRICE_URL)
	od_price_mapping = price_table_to_price_mapping(od_price_table)
	pre_od_price_table = get_price_table(LINUX_ON_DEMAND_PREVIOUS_GEN_PRICE_URL)
	pre_od_price_mapping = price_table_to_price_mapping(pre_od_price_table)
	# mapping from instance type to price info
	od_prices = od_price_mapping[region]
	pre_prices = pre_od_price_mapping[region]
	od_prices.update(pre_prices)
	# Get spot price lists
	p = subprocess.Popen(['ec2instancespricing.py', '--filter-region', region, '--type', 'spot' ,'--format', 'json'], stdout=subprocess.PIPE,stderr=subprocess.PIPE)
 	prices, err = p.communicate()
 	prices_json = json.loads(prices)
 	#pprint(prices_json['regions'])
 	# we assume there are 30 days in a month
 	month_hours = 30 * 24

	running_instances = {}
	spot_instances = {}
	spot_prices = {}
	for reservation in reservations:
		for instance in reservation.instances:
			if instance.state != "running":
				#print("Disqualifying instance {0}: not running\n".format( instance.id ))
				continue
			# get ondemand price
			az = instance.placement
			instance_type = instance.instance_type
			instance_id = instance.id
			
			if instance.spot_instance_request_id:
				# get spot price
				if (instance_type, az) not in spot_prices:
				 	for instanceType in prices_json['regions']:
						for each_type in instanceType['instanceTypes']:
							if each_type['type'] == instance_type:
								if each_type['utilization'] == 'spot':
									spot_prices[(instance_type, az)] = float(each_type['price'])
					pass

				spot_instances[ (instance_type, az ) ] = spot_instances.get( (instance_type, az ) , 0 ) + 1
			else:
				running_instances[ (instance_type, az ) ] = running_instances.get( (instance_type, az ) , 0 ) + 1
	# init excel file
	columns = [
		'Description',
		'Instance Count',
		'Instance Type',
		'Availability Zone',
		'Monthly Saving (USD)',
	]
	columns2 = [
		'Description', 
		'Instance Count', 
		'Instance Type', 
		'Availability Zone',
	]

	columns3 = [
		'Description', 
		'Instance Count', 
		'Instance Type', 
		'Availability Zone',
		'Potential RI Saving (USD)',
		'Potential Spot Saving (USD)',
	]

	wb = Workbook()
	dest_filename = 'aws_costs_'+region+'.xlsx'
	ws1 = wb.active
	ws1.title = 'Running Spot Instances'
	ws1.append(columns)
	ws2 = wb.create_sheet(title="Unused Reserved Instances")
	ws2.append(columns2)
	ws3 = wb.create_sheet(title='Unreserved Instances')
	ws3.append(columns3)
	# init pretty table
	
	table = PrettyTable(columns)
	table2 = PrettyTable(columns2)
	table3 = PrettyTable(columns3)
	for key in columns:
		table.align[key] = 'l'
	table.align['Monthly Saving'] = 'r'

	total_potential_spot_saving = 0
	# pprint( running_instances )
	# pprint(spot_instances)
	for (instance_type, az), value in spot_instances.iteritems():
		if instance_type in od_prices:
			od_price = float(od_prices[instance_type]['valueColumns'][0]['prices']['USD'])
			saving = (od_price - spot_prices[(instance_type, az)])*month_hours*value
			total_potential_spot_saving = total_potential_spot_saving + saving
			table.add_row(['Spot Instance', value, instance_type, az, saving])
			ws1.append(['Spot Instance', value, instance_type, az, saving])

	reserved_instances = {}
	reserved_instance_price = {}
	for reserved_instance in ec2_conn.get_all_reserved_instances():
		if reserved_instance.state != "active":
			#sys.stderr.write( "Excluding reserved instances %s: no longer active\n" % ( reserved_instance.id ) )
			continue
		else:
			az = reserved_instance.availability_zone
			instance_type = reserved_instance.instance_type
			if reserved_instance.recurring_charges:
				reserved_instance_price[(instance_type, az)] = float(reserved_instance.recurring_charges[0].amount)
			else:
				reserved_instance_price[(instance_type, az)] = 0
			reserved_instances[( instance_type, az) ] = reserved_instances.get ( (instance_type, az ), 0 )  + reserved_instance.instance_count

	# pprint( reserved_instances )

	# pprint (reserved_instance_price)
	# this dict will have a positive number if there are unused reservations
	# and negative number if an instance is on demand
	instance_diff = dict([(x, reserved_instances[x] - running_instances.get(x, 0 )) for x in reserved_instances])

	# instance_diff only has the keys that were present in reserved_instances. There's probably a cooler way to add a filtered dict here
	for placement_key in running_instances:
		if not placement_key in reserved_instances:
			instance_diff[placement_key] = -running_instances[placement_key]

	# pprint ( instance_diff )

	unused_reservations = dict((key,value) for key, value in instance_diff.iteritems() if value > 0)
	if unused_reservations == {}:
		print "Congratulations, you have no unused reservations"
	else:
		for unused_reservation in unused_reservations:
			ws2.append(['Unused Reservation',unused_reservations[ unused_reservation ], unused_reservation[0], unused_reservation[1]])
			table2.add_row(['Unused Reservation',unused_reservations[ unused_reservation ], unused_reservation[0], unused_reservation[1]])

	print ""

	unreserved_instances = dict((key,-value) for key, value in instance_diff.iteritems() if value < 0)
	if unreserved_instances == {}:
		print "Congratulations, you have no unreserved instances"
	else:
		for unreserved_instance in unreserved_instances:
			value = unreserved_instances[ unreserved_instance ]
			instance_type = unreserved_instance[0]
			az = unreserved_instance[1]
			# check for RI
			if (instance_type, az) in reserved_instance_price and instance_type in od_prices:
				od_price = float(od_prices[instance_type]['valueColumns'][0]['prices']['USD'])
				saving = (od_price - reserved_instance_price[(instance_type, az)])*month_hours*value
			else:
				saving = 'NA'

			# get spot price
			if (instance_type, az) not in spot_prices:
			 	for instanceType in prices_json['regions']:
					for each_type in instanceType['instanceTypes']:
						if each_type['type'] == instance_type:
							if each_type['utilization'] == 'spot':
								spot_prices[(instance_type, az)] = float(each_type['price'])
				pass

			# check for Spot
			if (instance_type, az) in spot_prices and instance_type in od_prices:
				od_price = float(od_prices[instance_type]['valueColumns'][0]['prices']['USD'])
				spot_saving = (od_price - spot_prices[(instance_type, az)])*month_hours*value
				total_potential_spot_saving = total_potential_spot_saving + spot_saving
			else:
				spot_saving = 'NA'
			ws3.append(['Unreserved Instance',value, instance_type, az, saving, spot_saving])
			table3.add_row(['Unreserved Instance',value, instance_type, az, saving, spot_saving])
			pass
	
	ws1.append([''])
	message = 'Total potential saving with Spot instances: USD {0}\n'.format(total_potential_spot_saving)
	ws1.append([message])
	if len(running_instances) > 0:
		qty_running_instances = reduce( lambda x, y: x+y, running_instances.values() )
	else:
		qty_running_instances = 0
	if len(reserved_instances) > 0:
		qty_reserved_instances = reduce( lambda x, y: x+y, reserved_instances.values() )
	else:
		qty_reserved_instances = 0
	if len(spot_instances) > 0:
		qty_spot_instances = reduce( lambda x, y: x+y, spot_instances.values() )
	else:
		qty_spot_instances = 0
	message = message + "\n(%s) running on-demand instances\n(%s) reservations\n(%s) spot instances" % ( qty_running_instances, qty_reserved_instances, qty_spot_instances)
	ws1.append(['{0} running on-demand instances'.format(qty_running_instances)])
	ws1.append(['{0} reservations'.format(qty_reserved_instances)])
	ws1.append(['{0} spot instances'.format(qty_spot_instances)])
	if (format == 'xlsx'):
		wb.save(dest_filename)
	else:
		print(table)
		print(table2)
		print(table3)
		print(message)